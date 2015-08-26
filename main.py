from kivy.app import App
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.widget import Widget
from kivy.properties import ObjectProperty
from kivy.uix.popup import Popup
from kivy.uix.image import Image
from kivy.core.window import Window

from openpyxl import load_workbook

import os

class StudentInfo(Widget):
    search_btn = ObjectProperty(None)
    sid_input = ObjectProperty(None)
    status_label = ObjectProperty(None)

    name_label = ObjectProperty(None)
    group_label = ObjectProperty(None)
    note_label = ObjectProperty(None)
    class_label = ObjectProperty(None)
    birthday_label = ObjectProperty(None)
    barcode_label = ObjectProperty(None)
    student_img = ObjectProperty(None)

    barcodes = {}

    ws = None

    def __init__(self, **kwargs):
        super(StudentInfo, self).__init__(**kwargs)
        self._keyboard = Window.request_keyboard(self._keyboard_closed, self)
        self._keyboard.bind(on_key_down=self._on_keyboard_down)

    def _on_keyboard_down(self, keyboard, keycode, text, modifiers):
        print keycode[1]
        if keycode[1] == 'enter':
            self.check_sid()

    def _keyboard_closed(self):
        self._keyboard.unbind(on_key_down=self._on_keyboard_down)
        self._keyboard = None

    def check_sid(self):
        if self.sid_input.text in self.barcodes.keys():
            self.show_record(self.barcodes[self.sid_input.text])
        else:
            self.status_label.text = 'Not Found !'

    def show_load(self):
        content = LoadDialog(load=self.load, cancel=self.dismiss_popup)
        self._popup = Popup(title="Load file", content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        count = self.read_workbook(os.path.join(path, filename[0]))
        if count < 0:
            self.status_label.text = 'Error in load file !'
        else:
            self.status_label.text = 'Read %d records !' % count
        self.dismiss_popup()

    def dismiss_popup(self):
        self._popup.dismiss()

    def read_workbook(self, file_path):
        try:
            work_book = load_workbook(file_path)
            self.ws = work_book[work_book.get_sheet_names()[0]]
            self.barcodes = {}
            i = 2
            c = self.ws.cell(row=i, column=1)
            while c.value:
                self.barcodes[str(c.value)] = i
                i += 1
                c = self.ws.cell(row=i, column=1)
            return len(self.barcodes.keys())
        except Exception, e:
            return -1

    def show_record(self, row):
        self.barcode_label.text = str(self.ws.cell(row=row, column=1).value)
        self.name_label.text = self.ws.cell(row=row, column=2).value
        self.class_label.text = str(self.ws.cell(row=row, column=3).value)
        self.birthday_label.text = str(self.ws.cell(row=row, column=4).value)
        self.group_label.text = self.ws.cell(row=row, column=5).value
        self.note_label.text = str(self.ws.cell(row=row, column=7).value)
        self.student_img.source = self.ws.cell(row=row, column=6).value
        self.student_img.reload()


class LoadDialog(FloatLayout):
    load = ObjectProperty(None)
    cancel = ObjectProperty(None)


class StudentApp(App):
    def build(self):
        return StudentInfo()


if __name__ == '__main__':
    StudentApp().run()
